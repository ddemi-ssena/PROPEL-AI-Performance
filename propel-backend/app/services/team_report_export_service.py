from __future__ import annotations

from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.schemas.team_report import TeamReportExportRequest


class TeamReportExportService:
    HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")
    SUBHEADER_FILL = PatternFill("solid", fgColor="DBEAFE")
    LIGHT_FILL = PatternFill("solid", fgColor="F8FAFC")
    WARN_FILL = PatternFill("solid", fgColor="FEF3C7")
    HIGH_FILL = PatternFill("solid", fgColor="FEE2E2")
    LOW_FILL = PatternFill("solid", fgColor="D1FAE5")
    WHITE_FONT = Font(color="FFFFFF", bold=True)
    BOLD_FONT = Font(bold=True)
    BORDER = Border(
        left=Side(style="thin", color="111827"),
        right=Side(style="thin", color="111827"),
        top=Side(style="thin", color="111827"),
        bottom=Side(style="thin", color="111827"),
    )

    @staticmethod
    def build_workbook(payload: TeamReportExportRequest) -> BytesIO:
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Ozet Rapor"
        TeamReportExportService._summary_sheet(summary, payload)
        TeamReportExportService._members_sheet(workbook.create_sheet("Takim Uyeleri Detay"), payload)
        TeamReportExportService._trend_sheet(workbook.create_sheet("12 Haftalik Trend"), payload)
        TeamReportExportService._risk_factors_sheet(workbook.create_sheet("Risk Faktorleri"), payload)
        TeamReportExportService._actions_sheet(workbook.create_sheet("Aksiyon Plani"), payload)

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream

    @staticmethod
    def _summary_sheet(sheet, payload: TeamReportExportRequest) -> None:
        sheet.merge_cells("A1:F1")
        sheet["A1"] = f"{payload.team.upper()} TAKIM ANALIZ RAPORU"
        sheet["A1"].fill = TeamReportExportService.HEADER_FILL
        sheet["A1"].font = Font(color="FFFFFF", bold=True, size=14)
        sheet["A1"].alignment = Alignment(horizontal="center")

        sheet["A2"] = "Tarih"
        sheet["B2"] = payload.report_date
        sheet["A3"] = "Rapor Turu"
        sheet["B3"] = payload.report_type

        row = 5
        TeamReportExportService._section_title(sheet, row, "TEMEL METRIKLER")
        row += 1
        for metric in payload.metrics:
            sheet.cell(row=row, column=1, value=metric.label)
            sheet.cell(row=row, column=2, value=metric.value)
            sheet.cell(row=row, column=2).alignment = Alignment(horizontal="right")
            sheet.cell(row=row, column=2).font = TeamReportExportService.BOLD_FONT
            TeamReportExportService._style_row(sheet, row, 1, 2)
            row += 1

        row += 2
        TeamReportExportService._section_title(sheet, row, "ANA SORUN")
        row += 1
        rows = [
            ("Konu", payload.main_issue_title),
            ("Aciklama", payload.main_issue_description),
            ("Ana Neden", payload.main_reason),
        ]
        for label, value in rows:
            sheet.cell(row=row, column=1, value=label)
            sheet.cell(row=row, column=2, value=value)
            sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            sheet.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            TeamReportExportService._style_row(sheet, row, 1, 6)
            if label == "Aciklama":
                sheet.row_dimensions[row].height = 54
            row += 1

        row += 2
        TeamReportExportService._section_title(sheet, row, "ONCELIKLI AKSIYONLAR")
        row += 1
        for index, action in enumerate(payload.actions[:5], start=1):
            sheet.cell(row=row, column=1, value=index)
            sheet.cell(row=row, column=2, value=action.title)
            sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            TeamReportExportService._style_row(sheet, row, 1, 6)
            row += 1

        TeamReportExportService._fit_columns(sheet, {"A": 24, "B": 28, "C": 18, "D": 18, "E": 18, "F": 18})

    @staticmethod
    def _members_sheet(sheet, payload: TeamReportExportRequest) -> None:
        members = sorted(
            payload.members,
            key=lambda member: member.risk_score if member.risk_score is not None else round(member.confidence * 10),
            reverse=True,
        )

        headers = ["No", "Isim Soyisim", "Rol", "Bolum", "Risk Skoru", "Risk Seviyesi", "Durum"]
        TeamReportExportService._write_header_row(sheet, 1, headers)
        for row_index, member in enumerate(members, start=2):
            risk_score = member.risk_score if member.risk_score is not None else round(member.confidence * 10)
            values = [
                row_index - 1,
                member.name,
                member.role or "-",
                member.department_code or TeamReportExportService._dataset_code(member.employee_id),
                risk_score,
                TeamReportExportService._risk_label(member.risk_level),
                member.status or TeamReportExportService._status_for_member(member.risk_level, risk_score),
            ]
            TeamReportExportService._write_table_row(sheet, row_index, values)

        if members:
            sheet.auto_filter.ref = f"A1:G{len(members) + 1}"
            sheet.conditional_formatting.add(
                f"E2:E{len(members) + 1}",
                DataBarRule(start_type="num", start_value=0, end_type="num", end_value=10, color="EF4444"),
            )

        extra_start = len(members) + 4
        sheet.merge_cells(start_row=extra_start, start_column=1, end_row=extra_start, end_column=7)
        section = sheet.cell(row=extra_start, column=1, value="EK BILGILER")
        section.fill = TeamReportExportService.HEADER_FILL
        section.font = TeamReportExportService.WHITE_FONT
        section.border = TeamReportExportService.BORDER

        extra_headers = ["Isim", "Motivasyon Skoru", "Is Tamamlama Orani", "Devamsizlik Gunu"]
        TeamReportExportService._write_header_row(sheet, extra_start + 1, extra_headers)
        for row_index, member in enumerate(members, start=extra_start + 2):
            values = [
                member.name,
                round(member.motivation_score, 1) if member.motivation_score is not None else "-",
                TeamReportExportService._percent_text(member.completion_rate),
                TeamReportExportService._days_text(member.absence_days),
            ]
            TeamReportExportService._write_table_row(sheet, row_index, values, end_column=4)
            if member.motivation_score is not None:
                sheet.cell(row=row_index, column=2).number_format = '0.0"/10"'

        if members:
            first = extra_start + 2
            last = extra_start + 1 + len(members)
            sheet.auto_filter.ref = f"A1:G{len(members) + 1}"
            sheet.conditional_formatting.add(
                f"B{first}:B{last}",
                ColorScaleRule(
                    start_type="num",
                    start_value=0,
                    start_color="EF4444",
                    mid_type="num",
                    mid_value=5,
                    mid_color="F59E0B",
                    end_type="num",
                    end_value=10,
                    end_color="10B981",
                ),
            )

        TeamReportExportService._fit_columns(sheet, {"A": 8, "B": 24, "C": 28, "D": 14, "E": 14, "F": 16, "G": 18})
        sheet.freeze_panes = "A2"

    @staticmethod
    def _trend_sheet(sheet, payload: TeamReportExportRequest) -> None:
        headers = ["Hafta", "Tarih", "Risk Skoru", "Motivasyon Ort.", "Kapasite Kullanimi"]
        TeamReportExportService._write_header_row(sheet, 1, headers)
        points = payload.trend[:12]
        for row_index, point in enumerate(points, start=2):
            capacity = point.capacity_usage
            capacity_text = TeamReportExportService._capacity_text(capacity)
            values = [
                point.period,
                point.date or "-",
                point.risk_score,
                point.motivation_avg if point.motivation_avg is not None else "-",
                capacity_text,
            ]
            TeamReportExportService._write_table_row(sheet, row_index, values)
            if capacity is not None and capacity > 100:
                sheet.cell(row=row_index, column=5).fill = TeamReportExportService.HIGH_FILL

        last_data_row = len(points) + 1
        change_row = last_data_row + 1
        sheet.cell(row=change_row, column=1, value="DEGISIM")
        sheet.cell(row=change_row, column=3, value=TeamReportExportService._change_text([point.risk_score for point in points]))
        sheet.cell(row=change_row, column=4, value=TeamReportExportService._change_text([
            point.motivation_avg for point in points if point.motivation_avg is not None
        ]))
        sheet.cell(row=change_row, column=5, value=TeamReportExportService._change_text([
            point.capacity_usage for point in points if point.capacity_usage is not None
        ]))
        TeamReportExportService._style_row(sheet, change_row, 1, 5)

        summary_row = change_row + 2
        TeamReportExportService._section_title(sheet, summary_row, "OZET ISTATISTIK")
        summary_headers = ["Metrik", "Ortalama", "Min", "Max", "Trend"]
        TeamReportExportService._write_header_row(sheet, summary_row + 1, summary_headers)
        summary_values = [
            ("Risk Skoru", [point.risk_score for point in points]),
            ("Motivasyon", [point.motivation_avg for point in points if point.motivation_avg is not None]),
            ("Kapasite", [point.capacity_usage for point in points if point.capacity_usage is not None]),
        ]
        for row_index, (label, values) in enumerate(summary_values, start=summary_row + 2):
            TeamReportExportService._write_table_row(
                sheet,
                row_index,
                [
                    label,
                    round(sum(values) / len(values), 2) if values else "-",
                    round(min(values), 2) if values else "-",
                    round(max(values), 2) if values else "-",
                    TeamReportExportService._trend_arrow(values),
                ],
                end_column=5,
            )

        eval_row = summary_row + 7
        TeamReportExportService._section_title(sheet, eval_row, "DEGERLENDIRME")
        evaluations = TeamReportExportService._trend_evaluations(points)
        for offset, text in enumerate(evaluations, start=1):
            sheet.cell(row=eval_row + offset, column=1, value=f"- {text}")
            sheet.merge_cells(start_row=eval_row + offset, start_column=1, end_row=eval_row + offset, end_column=5)
            TeamReportExportService._style_row(sheet, eval_row + offset, 1, 5)

        if points:
            sheet.auto_filter.ref = f"A1:E{last_data_row}"
            sheet.conditional_formatting.add(
                f"E2:E{last_data_row}",
                ColorScaleRule(
                    start_type="num",
                    start_value=80,
                    start_color="D1FAE5",
                    mid_type="num",
                    mid_value=100,
                    mid_color="FEF3C7",
                    end_type="num",
                    end_value=125,
                    end_color="EF4444",
                ),
            )
            chart = LineChart()
            chart.title = "12 Haftalik Risk ve Motivasyon Trendi"
            chart.y_axis.title = "Skor"
            chart.x_axis.title = "Hafta"
            data = Reference(sheet, min_col=3, max_col=4, min_row=1, max_row=last_data_row)
            categories = Reference(sheet, min_col=1, min_row=2, max_row=last_data_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.height = 9
            chart.width = 18
            sheet.add_chart(chart, "G2")

        TeamReportExportService._fit_columns(sheet, {"A": 14, "B": 16, "C": 14, "D": 18, "E": 20})
        sheet.freeze_panes = "A2"

    @staticmethod
    def _risk_factors_sheet(sheet, payload: TeamReportExportRequest) -> None:
        headers = ["No", "Risk Faktoru", "Etki Seviyesi", "Olasilik", "Oncelik"]
        TeamReportExportService._write_header_row(sheet, 1, headers)
        for row_index, factor in enumerate(payload.risk_factors, start=2):
            severity = factor.impact_level or factor.severity
            probability = factor.probability if factor.probability is not None else TeamReportExportService._probability_for_factor(factor.count, severity)
            values = [
                row_index - 1,
                factor.name,
                TeamReportExportService._risk_label(severity),
                f"%{probability}",
                factor.priority or TeamReportExportService._priority_for_factor(severity, probability, row_index - 1),
            ]
            TeamReportExportService._write_table_row(sheet, row_index, values)

        if payload.risk_factors:
            sheet.auto_filter.ref = f"A1:E{len(payload.risk_factors) + 1}"

        detail_row = len(payload.risk_factors) + 4
        TeamReportExportService._section_title(sheet, detail_row, "DETAYLI ACIKLAMALAR")
        row = detail_row + 1
        for index, factor in enumerate(payload.risk_factors, start=1):
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            title = sheet.cell(row=row, column=1, value=f"{index}. {factor.name.upper()}")
            title.fill = TeamReportExportService.SUBHEADER_FILL
            title.font = TeamReportExportService.BOLD_FONT
            title.border = TeamReportExportService.BORDER
            row += 1
            detail_lines = [
                ("Mevcut Durum", factor.current_state or factor.note or "Risk sinyali izleniyor"),
                ("Hedef", factor.target_state or "Risk seviyesini kabul edilebilir banda indirmek"),
                ("Acik (Gap)", factor.gap or "-"),
                ("Etkilenen Kisi", factor.affected_people or f"{factor.count} kisi"),
                ("Beklenen Sonuc", factor.expected_result or "Performans dususu ve takip ihtiyaci"),
            ]
            for label, value in detail_lines:
                sheet.cell(row=row, column=1, value=f"- {label}:")
                sheet.cell(row=row, column=2, value=value)
                sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
                TeamReportExportService._style_row(sheet, row, 1, 5)
                row += 1
            row += 1

        TeamReportExportService._fit_columns(sheet, {"A": 8, "B": 36, "C": 16, "D": 14, "E": 16})
        sheet.freeze_panes = "A2"

    @staticmethod
    def _actions_sheet(sheet, payload: TeamReportExportRequest) -> None:
        headers = ["No", "Aksiyon", "Sorumlu", "Hedef Tarih", "Oncelik", "Durum"]
        TeamReportExportService._write_header_row(sheet, 1, headers)
        for row_index, action in enumerate(payload.actions, start=2):
            values = [
                row_index - 1,
                action.title,
                action.owner or "Takim lideri",
                action.target_date or action.timeframe or "Bu hafta",
                action.priority or TeamReportExportService._priority_for_action(row_index - 1),
                action.status or "\u23f3 Bekle",
            ]
            TeamReportExportService._write_table_row(sheet, row_index, values)

        if payload.actions:
            sheet.auto_filter.ref = f"A1:F{len(payload.actions) + 1}"

        detail_row = len(payload.actions) + 4
        TeamReportExportService._section_title(sheet, detail_row, "AKSIYON DETAYLARI")
        row = detail_row + 1
        for index, action in enumerate(payload.actions, start=1):
            details = [
                f"{index}. {action.title}",
                f"Neden: {action.reason or '-'}",
                f"Beklenen Etki: {action.expected_impact or '-'}",
            ]
            for text in details:
                sheet.cell(row=row, column=1, value=text)
                sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                TeamReportExportService._style_row(sheet, row, 1, 6)
                row += 1
            row += 1

        topic_row = row + 1
        TeamReportExportService._section_title(sheet, topic_row, "KONUSULACAK KONULAR")
        for offset, point in enumerate(payload.talking_points, start=1):
            sheet.cell(row=topic_row + offset, column=1, value=f"\u2610 {point}")
            sheet.merge_cells(start_row=topic_row + offset, start_column=1, end_row=topic_row + offset, end_column=6)
            TeamReportExportService._style_row(sheet, topic_row + offset, 1, 6)

        TeamReportExportService._fit_columns(sheet, {"A": 8, "B": 38, "C": 18, "D": 16, "E": 14, "F": 14})
        sheet.freeze_panes = "A2"

    @staticmethod
    def _table(sheet, headers: list[str], rows: Iterable[list[object]]) -> None:
        TeamReportExportService._write_header_row(sheet, 1, headers)

        for row_index, row_values in enumerate(rows, start=2):
            TeamReportExportService._write_table_row(sheet, row_index, row_values)

        TeamReportExportService._fit_columns(sheet)
        sheet.freeze_panes = "A2"

    @staticmethod
    def _write_header_row(sheet, row: int, headers: list[str]) -> None:
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=column, value=header)
            cell.fill = TeamReportExportService.HEADER_FILL
            cell.font = TeamReportExportService.WHITE_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = TeamReportExportService.BORDER

    @staticmethod
    def _write_table_row(sheet, row: int, values: Iterable[object], end_column: int | None = None) -> None:
        fill = TeamReportExportService.LIGHT_FILL if row % 2 == 0 else None
        values_list = list(values)
        max_column = end_column or len(values_list)
        for column in range(1, max_column + 1):
            value = values_list[column - 1] if column <= len(values_list) else None
            cell = sheet.cell(row=row, column=column, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = TeamReportExportService.BORDER
            if fill:
                cell.fill = fill
            if isinstance(value, str):
                normalized = value.lower()
                if "yuksek" in normalized or "\U0001f534" in value:
                    cell.fill = TeamReportExportService.HIGH_FILL
                elif "orta" in normalized or "\U0001f7e1" in value:
                    cell.fill = TeamReportExportService.WARN_FILL
                elif "dusuk" in normalized or "\U0001f7e2" in value:
                    cell.fill = TeamReportExportService.LOW_FILL

    @staticmethod
    def _risk_label(value: str) -> str:
        normalized = value.lower()
        if "yuksek" in normalized or normalized == "high":
            return "\U0001f534 Yuksek"
        if "orta" in normalized or normalized == "medium":
            return "\U0001f7e1 Orta"
        return "\U0001f7e2 Dusuk"

    @staticmethod
    def _status_for_member(risk_level: str, risk_score: int) -> str:
        normalized = risk_level.lower()
        if "yuksek" in normalized or normalized == "high":
            return "Acil Mudahale" if risk_score >= 9 else "Takip Gerekli"
        if "orta" in normalized or normalized == "medium":
            return "Izleniyor" if risk_score >= 6 else "Stabil"
        return "Stabil"

    @staticmethod
    def _dataset_code(employee_id: int) -> str:
        return f"SE-{employee_id:03d}"

    @staticmethod
    def _score_text(value: float | None, suffix: str) -> str:
        if value is None:
            return "-"
        return f"{round(value, 1)}{suffix}"

    @staticmethod
    def _percent_text(value: float | None) -> str:
        if value is None:
            return "-"
        normalized = value * 100 if 0 <= value <= 1 else value
        return f"%{round(normalized)}"

    @staticmethod
    def _days_text(value: float | None) -> str:
        if value is None:
            return "-"
        return f"{round(value, 1):g} gun"

    @staticmethod
    def _capacity_text(value: float | None) -> str:
        if value is None:
            return "-"
        suffix = " \U0001f534" if value > 100 else ""
        return f"%{round(value):g}{suffix}"

    @staticmethod
    def _change_text(values: list[float | None]) -> str:
        numeric = [float(value) for value in values if value is not None]
        if len(numeric) < 2 or numeric[0] == 0:
            return "-"
        change = ((numeric[-1] - numeric[0]) / abs(numeric[0])) * 100
        sign = "+" if change > 0 else ""
        return f"{sign}{round(change)}%"

    @staticmethod
    def _trend_arrow(values: list[float | None]) -> str:
        numeric = [float(value) for value in values if value is not None]
        if len(numeric) < 2:
            return "\u2192"
        delta = numeric[-1] - numeric[0]
        if delta > 0:
            return "\u2191"
        if delta < 0:
            return "\u2193"
        return "\u2192"

    @staticmethod
    def _trend_evaluations(points) -> list[str]:
        risk_change = TeamReportExportService._change_text([point.risk_score for point in points])
        motivation_change = TeamReportExportService._change_text([
            point.motivation_avg for point in points if point.motivation_avg is not None
        ])
        capacity_values = [point.capacity_usage for point in points if point.capacity_usage is not None]
        latest_capacity = capacity_values[-1] if capacity_values else None
        capacity_change = TeamReportExportService._change_text(capacity_values)
        items = []
        if risk_change != "-":
            items.append(f"Risk skorunda 12 haftada {risk_change} degisim")
        if motivation_change != "-":
            items.append(f"Motivasyon skorunda 12 haftada {motivation_change} degisim")
        if latest_capacity is not None:
            status = "kritik seviyede" if latest_capacity > 100 else "kontrol altinda"
            items.append(f"Sprint kapasitesi {status} (%{round(latest_capacity):g}, degisim {capacity_change})")
        return items or ["Trend degerlendirmesi icin yeterli veri yok"]

    @staticmethod
    def _probability_for_factor(count: int, severity: str) -> int:
        normalized = severity.lower()
        if "high" in normalized or "yuksek" in normalized:
            return min(100, max(85, 70 + (count * 5)))
        if "medium" in normalized or "orta" in normalized:
            return min(85, max(55, 45 + (count * 5)))
        return min(60, max(30, 25 + (count * 5)))

    @staticmethod
    def _priority_for_factor(severity: str, probability: int, index: int) -> str:
        normalized = severity.lower()
        if ("high" in normalized or "yuksek" in normalized) and probability >= 90:
            return "P0 - Acil"
        if "high" in normalized or "yuksek" in normalized:
            return "P1 - Yuksek"
        if "medium" in normalized or "orta" in normalized:
            return "P2 - Orta"
        return "P3 - Dusuk"

    @staticmethod
    def _priority_for_action(index: int) -> str:
        if index <= 2:
            return "P0"
        if index <= 4:
            return "P1"
        return "P2"

    @staticmethod
    def _section_title(sheet, row: int, title: str) -> None:
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = sheet.cell(row=row, column=1, value=title)
        cell.fill = TeamReportExportService.HEADER_FILL
        cell.font = TeamReportExportService.WHITE_FONT
        cell.border = TeamReportExportService.BORDER

    @staticmethod
    def _style_row(sheet, row: int, start_column: int, end_column: int) -> None:
        for column in range(start_column, end_column + 1):
            cell = sheet.cell(row=row, column=column)
            cell.border = TeamReportExportService.BORDER
            if column == start_column:
                cell.fill = TeamReportExportService.SUBHEADER_FILL
                cell.font = TeamReportExportService.BOLD_FONT

    @staticmethod
    def _fit_columns(sheet, fixed: dict[str, int] | None = None) -> None:
        fixed = fixed or {}
        for column_cells in sheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            if column_letter in fixed:
                sheet.column_dimensions[column_letter].width = fixed[column_letter]
                continue
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_letter].width = min(max(max_length + 3, 14), 42)
